from openpyxl import load_workbook
from asgiref.sync import sync_to_async
from .ozon_parser import *
from .models import (
    OzonIndexerReport, 
    OzonNmidToBeReported, 
    OzonIndexerReportData, 
    DateCreateReports,
    User,
)

async def ozon_parser(date_create_reports, *nmids_and_users):
    nmids = list(map(lambda x: int(x[0]), nmids_and_users))
    users_id = list(map(lambda x: x[1], nmids_and_users))
    parser = OzonIndexerManager(*nmids)
    await parser.run()

    for i, result in enumerate(parser.results):
        if type(result['data']) is str:
            print(f"'{nmids[i]}' it's not defined")
        elif result is not None:
            report = await OzonIndexerReport.objects.acreate(
                nmid=result['nmid'], 
                user=User(users_id[i]), 
                date_create_reports=DateCreateReports(date_create_reports)
            )

            for data in result['data']:
                await OzonIndexerReportData.objects.acreate(
                    **data,
                    report=report
                )

            report.ready = True
            await report.asave()
        else:
            print(f"'{nmids[i]}' it's not defined")


async def create_nmid_to_report(nmid, user_id):
    async with OzonParser() as parser:
        item = OzonProductItem(await parser.get_full_info_from_api_product(nmid))
        title = item.title

        await OzonNmidToBeReported.objects.aget_or_create(
            nmid=nmid,
            name=title,
            user=User(pk=user_id),
        )


class FileOperator:

    def iterate_nmids(self, filepath: str):
        book = load_workbook(filepath)
        sheet = book.active
        nmids = set()
        for cell in sheet.iter_rows(values_only=True):
            nmid = cell[0]
            nmids.add(nmid)

        return nmids


    def create_report_buffer(self, report_id):

        buffer = io.BytesIO()
        book = xlsxwriter.Workbook(buffer)
        sheet = book.add_worksheet()

        # создаем строку с названием всех колонок
        columns = [
            'priority_cat', 'keywords', 'frequency', 'req_depth', 'existence', 'place', 'spot_req_depth', 'ad_spots',
            'ad_place'
        ]

        for i in range(len(columns)):
            sheet.write(0, i, columns[i])

        # собираем данные из бд, итерируем, заносим в лист
        data = OzonIndexerReportsData.objects.filter(report_id=report_id)
        row_counter = 1
        for query in data:
            sheet.write(row_counter, 0, query.priority_cat)
            sheet.write(row_counter, 1, query.keywords)
            sheet.write(row_counter, 2, query.frequency)
            sheet.write(row_counter, 3, query.req_depth)
            sheet.write(row_counter, 4, query.existence)
            sheet.write(row_counter, 5, query.place)
            sheet.write(row_counter, 6, query.spot_req_depth)
            sheet.write(row_counter, 7, query.ad_spots)
            sheet.write(row_counter, 8, query.ad_place)
            row_counter += 1

        book.close()
        buffer.seek(0)
        return buffer