import asyncio
import time, re, xlsxwriter, io, os, json, base64, random, fake_useragent
from pymorphy3 import MorphAnalyzer
from openpyxl import load_workbook
from . import models
from pathlib import Path
from datetime import datetime, timedelta
from httpx import AsyncClient
from asgiref.sync import async_to_sync


HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Linux; Android 14) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.6099.230 Mobile Safari/537.36',
    'Cookie': '_wbauid=6692713411689684383; ___wbu=49933369-b549-4f32-9316-ccaf78e81a72.1689684385; x-supplier-id-external=9b712861-ed95-4f81-bb88-e514331d222d; external-locale=ru; wbx-validation-key=4e088847-45df-4e53-ad99-5a32e1e2be8f; __zzatw-wb=MDA0dBA=Fz2+aQ==; WILDAUTHNEW_V3=0CDEC72CED216FC25C8D2818FDC293E9C76061080331E7A85669CA48F3C5F4AC1C46F14D4648AFCA8EA18915632D2A59071E7819D61CC2AF1682E9C8C31E05466170DA1D13366F0F1D30A4936A1464E955F7BC9F783B7E6C72E2B55D49CC81376320E8BAB11AFA29B77E89ED9347099C29BFEA5DA6AC689DB67AFC300ECE0B45A3AA76BDEB6442DDF9CAA68FF1F059BF8CAD07FE89C9A2F298BA81A648727C19F9016807424E3977BF400564B18EB8B787B5B03AE77D05958F40D49B1BAFB4E284A44DA4EF6E46E1F2398DC0331A67F5F59E66EB605EFE6E5588CF96EC6B28AE7BF6BCC846CC7885C8B9FE4D5A78AFE5ABB4A04BD2552DE301D76BCF54F8BE5CDC1DF2D4A0B0EE7237064DA231DE9BCF52224B974A7CB1350892F819F0B63B77221D4B3C0743081E934B781C587E74720FE56792; um=uid%3Dw7TDssOkw7PCu8KywrPCtsK1wrPCtsK0wrY%253d%3Aproc%3D89%3Aehash%3D; ___wbs=9af672e8-a0e6-41ec-98f0-3dc32241de85.1706016683; WBToken=AuW6zg-mlP7aDKbokdwMVI51I0QcCgwCH7VictwXypLqQOr2GzkBjFuoHEKnIIEqVrjgjlTkO2xyF7vRXGPJD2ouh09M_JruRw2F-3vLjTmLJEB-MSY3N0KfQOH9VpQrBcLn9g; cfidsw-wb=ImvGc0q+feUO3xzirHvDCKbWb7EtZswUxoSFXVPV27GQxfA0U8lFmGO8vIMsjkx0oXbqEi1lFeFpfzqVKmiXh9VfCI+uvJjkwPZgODVaNHw3zy/SWquDq84i/OGPZpl1fWa9BAmuRRBY41S7w0xZYHQaa9qjbsCc/qkUPw=='
}

def check_int(el):
    try:
        num = int(el)
        return num
    except:
        return None

# приводит каждое слово в переданном тексте к начальной форме
def normalize_text(text: str) -> str:
    morph = MorphAnalyzer()
    words = re.sub(r'[^\w\s]', '', text).split(' ')
    normalized_words = []
    for word in words:
        normalized_word = morph.parse(word)[0].normal_form
        normalized_words.append(normalized_word)

    return ' '.join(normalized_words)

def create_quick_indexation_report(report_id, nmid):

    report = models.IndexerReport.objects.get(id=report_id)
    indexer = Indexer(nmid)
    requests_data = list(models.Request.objects.all())
    async_to_sync(indexer.search_common)(requests_data)
    for query in indexer.iterate_resulted_queries():
        models.IndexerReportData.objects.create(
            priority_cat=query.get('top_category'),
            keywords=query.get('keywords'),
            frequency=query.get('frequency'),
            req_depth=query.get('req_depth'),
            existence=query.get('existence'),
            place=query.get('place'),
            spot_req_depth=query.get('spot_req_depth'),
            ad_place=query.get('ad_place'),
            report=report,
            product_id=nmid,
            quick_indexation=report.quick_indexation
        )

    report.ready = True
    now = datetime.now()
    report.time_of_readiness = now
    report.save()



# записывает все данные по каждому отчету в IndexerReportData
def createReportData(report_id, nmid):
    report = models.IndexerReport.objects.get(id=report_id)
    nmid_obj = models.NmidToBeReported.objects.all().filter(nmid=nmid)[0]
    indexer = Indexer(nmid)
    phrase = nmid_obj.phrase

    if phrase:
    #     print('no phrase data collecting started')
    #     requests_data = list(models.Request.objects.all())
    #     async_to_sync(indexer.search_common)(requests_data)
    #     for query in indexer.iterate_resulted_queries():
    #         models.IndexerReportData.objects.create(
    #             priority_cat=query.get('top_category'),
    #             keywords=query.get('keywords'),
    #             frequency=query.get('frequency'),
    #             req_depth=query.get('req_depth'),
    #             existence=query.get('existence'),
    #             place=query.get('place'),
    #             spot_req_depth=query.get('spot_req_depth'),
    #             ad_place=query.get('ad_place'),
    #             report=report,
    #             product_id=nmid
    #         )
    #
    # else:
        standard_queries = models.SeoCollectorPhraseData.objects.all().filter(phrase=phrase, standard=True)
        for query in indexer.iterate_standard_queries(standard_queries):
            models.IndexerReportData.objects.create(
                priority_cat=query.get('top_category'),
                keywords=query.get('keywords'),
                frequency=query.get('frequency'),
                req_depth=query.get('req_depth'),
                existence=query.get('existence'),
                place=query.get('place'),
                spot_req_depth=query.get('spot_req_depth'),
                ad_place=query.get('ad_place'),
                report=report,
                product_id=nmid, quick_indexation=report.quick_indexation
            )

    report.ready = True
    report.save()


class ProxyOperator:

    proxies = [
        'http://dsOi3l:tCdS6WFAPJ@46.8.154.154:1050',
        'http://dsOi3l:tCdS6WFAPJ@109.248.13.227:1050',
        'http://dsOi3l:tCdS6WFAPJ@188.130.187.143:1050',
        'http://dsOi3l:tCdS6WFAPJ@109.248.204.253:1050',
        'http://dsOi3l:tCdS6WFAPJ@46.8.192.30:1050',
    ]

    def get_random_proxy(self):
        proxy = random.choice(self.proxies)
        proxies = {
            'http://': proxy,
            'https://': proxy
        }


class NmidContextOperator:

    def __init__(self, requests: dict, reports):
        self.requests = requests
        self.reports = reports
        self.__create_context()
        self.requests = dict(sorted(self.requests.items()))



    def __iterate_report_data(self, report_data):

        for request in self.requests:
            if request in report_data:
                row_data = report_data[request]
                self.requests[request]['data'].append(
                    {'place': row_data.place, 'req_depth': row_data.req_depth,
                     'frequency': row_data.frequency, 'date': row_data.date.strftime("%m/%d/%y")}
                )
            else:
                self.requests[request]['data'].append(
                    {'place': None, 'req_depth': None,
                    'frequency': None, 'date': None}
                )



    def __create_context(self):
        for report in self.reports:
            report_data_queryset = models.IndexerReportData.objects.all().filter(report=report)
            report_data = dict(zip(
                [data.keywords for data in report_data_queryset],
                list(report_data_queryset)
            ))

            self.__iterate_report_data(report_data)


# класс для работы с файлами
# открытие, чтение, запись, вся хуйня
class FileOperator:
    # для работы с csv файлами используется встроенная библиотека csv
    # для работы с excel файлами требуется установка openpyexcel

    # метод для получения nmid товаров из эксель файла
    def iterate_excel_file(self, filepath: str):
        book = load_workbook(filepath)
        sheet = book.active
        nmids = set()
        for cell in sheet.iter_rows(values_only=True):
            nmid = cell[0]
            nmids.add(nmid)

        return nmids

    # применение морфологического анализатора для csv файлика
    # содержащего топ запросы за х времени
    # def rewrite_top_requests(self):
    #     reader = csv.reader('requests.csv')
    #     with open('normalized_requests.csv', 'w', encoding='cp1251', newline='') as file:
    #         writer = csv.writer(file)
    #
    #         for query in reader:
    #             keywords = query[0]
    #             frequency = query[1]
    #             keywords_normalized = normalize_text(keywords)
    #             normalized_query = [
    #                 keywords_normalized, frequency
    #             ]
    #
    #             writer.writerow(normalized_query)

    # создает файл с отчетом по определенному товару
    # выгружает данные в xlsx формате
    def create_indexer_report_buffer(self, report_id):

        # инициализируем все нужные инструменты
        buffer = io.BytesIO()
        book = xlsxwriter.Workbook(buffer)
        sheet = book.add_worksheet()

        # создаем строку с названием всех колонок
        columns = [
            'priority_cat', 'keywords', 'frequency', 'req_depth', 'existence', 'place', 'spot_req_depth', 'ad_spots',
            'ad_place'
        ]

        for i in range(len(columns)):
            sheet.write(0, i, columns[i])

        # собираем данные из бд, итерируем, заносим в лист
        report = models.IndexerReport.objects.all().filter(id=report_id).first()
        data = models.IndexerReportData.objects.all().filter(report=report)
        row_counter = 1
        for query in data:
            sheet.write(row_counter, 0, query.priority_cat)
            sheet.write(row_counter, 1, query.keywords)
            sheet.write(row_counter, 2, query.frequency)
            sheet.write(row_counter, 3, query.req_depth)
            sheet.write(row_counter, 4, query.existence)
            sheet.write(row_counter, 5, query.place)
            sheet.write(row_counter, 6, query.spot_req_depth)
            sheet.write(row_counter, 7, query.ad_spots)
            sheet.write(row_counter, 8, query.ad_place)
            row_counter += 1

        book.close()
        buffer.seek(0)
        return buffer

    def create_seo_report_buffer(self, phrase_id):
        buffer = io.BytesIO()
        book = xlsxwriter.Workbook(buffer)
        sheet = book.add_worksheet()

        columns = [
            'Ключевые слова', "Частотность", "Глубина", "Категория"
        ]

        for i in range(len(columns)):
            sheet.write(0, i, columns[i])
        phrase_obj = models.SeoCollectorPhrase.objects.all().filter(id=phrase_id).first()
        data = models.SeoCollectorPhraseData.objects.all().filter(phrase=phrase_obj, standard=True)
        print(len(data))
        row_counter = 1

        for query in data:

            sheet.write(row_counter, 0, query.query)
            sheet.write(row_counter, 1, query.frequency)
            sheet.write(row_counter, 2, query.depth)
            sheet.write(row_counter, 3, query.priority_cat)
            row_counter += 1

        book.close()
        buffer.seek(0)
        return buffer





    # # приводит все слова всех запросов к начальной форме
    # def normalizeRequestsFile(self):
    #     normalizedRequests = []
    #
    #     with open(self.requests_file_path) as requestsJsonFile:
    #         notNormalizedRequests = json.load(requestsJsonFile)
    #         counter = 1
    #         for query in notNormalizedRequests:
    #             print(counter)
    #             counter += 1
    #             keywords = query.get('keywords')
    #             normalizedKeywords = normalize_text(keywords)
    #             normalizedRequests.append(
    #                 {
    #                     'keywords': keywords,
    #                     'frequency': query.get('frequency'),
    #                     'normalizedKeywords': normalizedKeywords
    #                 }
    #             )
    #
    #     with open(self.requests_file_path, 'w', encoding='utf-8') as requestsJsonFile:
    #         data = json.dumps(normalizedRequests, indent=4, ensure_ascii=False)
    #         requestsJsonFile.write(data)

# класс предназначен для работы с ссылками
# содержит набор статичных ссылок и шаблонов для использования api вайлдберрис
class URLOperator:

    query_categories_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&filters=xsubject&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=filters&spp=0'
    subject_base_url = 'https://static-basket-01.wb.ru/vol0/data/subject-base.json'
    card_url_template = 'https://basket-replace_me.wb.ru/vol{}/part{}/{}/info/ru/card.json'
    ad_url_template = 'https://catalog-ads.wildberries.ru/api/v5/search?keyword={}'
    filtered_by_brand_id_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&fbrand={brand_id}&page=replace_me&query={query}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'
    any_query_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&page=replace_me&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=catalog&sort=popular&spp=0&suppressSpellcheck=false'
    any_query_req_depth_url_template = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&appType=1&curr=rub&dest=123586150&filters=xsubject&query={}&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=filters&spp=0'
    nmid_detail_url_template = 'https://card.wb.ru/cards/detail?appType=1&curr=rub&dest=123586150&regions=80,38,4,64,83,33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&spp=0&nm={}'


    card_url_template_ranges = {
        '01': (0, 143), '02': (144, 287), '03': (288, 431),
        '04': (432, 719), "05":(720, 1007), "06":(1008, 1061),
        "07": (1062, 1115), '08': (1116, 1169), '09': (1170, 1313),
        '10': (1314, 1601), "11": (1602, 1655), '12': (1656, 1919)
    }


    # метод формирует ссылку для сбора данных с карточки определенного товара
    def create_card_url(self, product_id: int) -> str:

        nmid = str(product_id)
        lenght = len(nmid)
        vol = lenght-5
        part = lenght-3

        pre_url = self.card_url_template.format(nmid[:vol], nmid[:part], nmid)

        for backet_number in self.card_url_template_ranges:
            start = self.card_url_template_ranges[backet_number][0]
            end = self.card_url_template_ranges[backet_number][1]
            if start <= int(product_id) // 100000 <= end:
                return pre_url.replace('replace_me', backet_number)

        return pre_url.replace('replace_me', '13')


    # создание ссылки для сбора информации о рекламе
    def create_ad_url(self, query: str) -> str:
        return self.ad_url_template.format(
            query.strip()
        )



    # это возможно чуть позже будет переделываться или вообще удалено будет
    # здесь создается ссылка для сбора данных по продавцу
    # ссылка используется для определения существования товара по определенному запросу
    def create_filtered_by_brand_url(self, query, brand_id):
        keywords = query.replace(' ', '%20')
        return self.filtered_by_brand_id_url_template.format(
            brand_id=brand_id,
            query=keywords.strip()
        ).replace('replace_me', '1')


    # создание ссылки для получения глубины выдачи
    def create_query_req_depth_url(self, query):
        return self.any_query_req_depth_url_template.format(
            query.strip()
        )



    # создает ссылку для получения информации по опреденному запросу
    # без каких либо фильтров или орграничений
    def create_query_url(self, query):
        return self.any_query_url_template.format(
            query.strip()
        )

    # создает ссылку для получения инфы от апи по категориям по определенному запросу в выдаче
    # вызывается в том случае, если топ категория по определенному запросу не была получена
    def create_query_categories_url(self, query):
        # print(f'{query} query text')
        # print(f"{self.query_categories_url_template.format(query.strip())} url text")
        return self.query_categories_url_template.format(
            query.strip()
        )

    def create_nmid_detail_url(self, nmid):
        return self.nmid_detail_url_template.format(nmid)


# класс DataCollector отправляет запросы на ссылки, которые были сгенерированы с помощью класса URLOperator
# получает и парсит данные от апи, а также обрабатывает возможные ошибки
class DataCollector:

    proxy_operator = ProxyOperator()
    user_agent = fake_useragent.UserAgent().random


    # выдергивает описание товара из пришедшей линки
    # линку предварительно сгенерировал юрл оператор в вызове методе create_card_url
    async def get_card_info(self, card_url):
        proxies = self.proxy_operator.get_random_proxy()
        async with AsyncClient(proxies=proxies) as client:
            resp = await client.get(card_url, timeout=None, headers=HEADERS)
            resp = resp.json()
            info = ''
            try:
                desc = resp.get('description', '')
                info += desc + ' '
                grouped_options_info = resp.get('grouped_options', {})
                options_info = resp.get('options', {})
                grouped_options = grouped_options_info[0]['options']
                for option in grouped_options:
                    info += option['value'] + ' '

                for key in options_info:
                    info += option['value'] + ' '

                return info

            except Exception as e:
                print(e, card_url)
                return ''

    # сбор id бренда
    async def get_brand_id(self, detail_url):
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(detail_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                return resp.get('data').get('products')[0].get('brandId')
            except Exception as e:
                print(f'error at get_brand_id {detail_url} {e}')


    # получение глубины запроса
    async def get_req_depth(self, query_depth_url, try_counter=0):
        if try_counter <= 5:
            proxies = self.proxy_operator.get_random_proxy()
            async with AsyncClient(proxies=proxies) as client:
                try:
                    resp = await client.get(query_depth_url, timeout=None, headers={'user-agent': self.user_agent})
                    resp = resp.json()
                    return resp.get('data').get('total')
                except Exception as e:
                    new_try_counter = try_counter + 1
                    await asyncio.sleep(15)
                    result = await self.get_req_depth(query_depth_url, new_try_counter)
                    return result
        else:
            print(f'error at get_req_depth {query_depth_url}')
            return 0

    # получение id товаров по определенному брэнду
    # возможно тоже будет переделано или удалено
    # nmid товара при определенном запросе нужны для проверки наличия по запросу
    async def get_query_by_brand(self, query_by_brand_url, try_counter=0):
        if try_counter <= 5:

            proxies = self.proxy_operator.get_random_proxy()

            async with AsyncClient(proxies=proxies) as client:
                counter = 1
                ids = set()
                query_by_brand_url.replace('replace_me', str(counter))
                query_by_brand_url = query_by_brand_url.replace(f'page={counter - 1}', f'page={counter}')
                try:
                    resp = await client.get(query_by_brand_url, timeout=None, headers=HEADERS)
                    resp = resp.json()
                    products = resp.get('data').get('products')

                    if not products:
                        return ids
                    else:
                        for product in products:
                            ids.add(str(product.get('id', '')))

                except Exception:
                    return self.get_query_by_brand(query_by_brand_url, try_counter+1)

        else:
            return set()


    # возвращает все nmid товаров с первых 10 (если есть) страниц
    async def get_query(self, query_url):
        counter = 1
        ids = []
        query_url = query_url.replace('replace_me', str(counter))
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            while counter <= 10:
                try:
                    query_url = query_url.replace(f'page={counter-1}', f'page={counter}')
                    resp = await client.get(query_url, timeout=None, headers=HEADERS)
                    resp = resp.json()
                    products = resp.get('data').get('products')
                    if not products:
                        return ids

                    for product in products:
                        ids.append(str(product.get('id', '')))

                    counter += 1

                except Exception as e:
                    time.sleep(20)
                    print(f'error at get_query {query_url} {e}')

            return ids

    # возвращает все рекламные nmid товаров
    async def get_ad(self, ad_url):
        ad_ids = []
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(ad_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                adverts = resp.get('adverts')
                if adverts:
                    for product in adverts:
                        ad_ids.append(str(product.get('id', '')))

                return ad_ids

            except Exception as e:
                print(f'error at get_ad {ad_url} {e}')
                return []

    # возвращает топ категорию из апи ответа по рекламе
    async def get_top_category(self, ad_info_url: str):
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(ad_info_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                priroty_cats = resp.get('prioritySubjects')
                if priroty_cats:
                    return priroty_cats[0]

                return ''

            except Exception as e:
                print(f'error at get_top_category {ad_info_url} {e}')
                return ''

    # возвращает все категории, которые есть на вб
    async def get_subject_base(self, subject_base_url: str):

        categories = {}
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                 resp = await client.get(subject_base_url, timeout=None, headers=HEADERS)
                 resp = resp.json()
                 for parent_category in resp:
                     for child_category in parent_category['childs']:
                         categories.update(
                             {
                                 child_category['id']: child_category['name']
                             }
                         )
                 return categories

            except Exception as e:
                print(f'error at get_subject_base {subject_base_url} {e}')
                return {}


    # если нет топ категории по рекламе(запрос уебанский), то этот метод возвращает категорию
    # у которой наибольшее число товаров по опреденному запросу
    async def get_most_category(self, query_categories_url):
        category = ''
        count = 0
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(query_categories_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                for key in resp.get('data').get('filters'):
                    for item in key.get('items'):
                        if item.get('count') > count:
                            category = item.get('name')
                            count = item.get('count')

                return category

            except Exception as e:
                print(f'error at  get_query_most_category {query_categories_url} {e}')
                return ''



    # возвращает бренд и имя товара
    # для более подробного описания текстовой части карточки
    async def get_brand_and_name(self, detail_url):
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(detail_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                product = resp.get('data').get('products')[0]
                name = product.get('name')
                brand = product.get('brand')
                return [name, brand]

            except Exception as e:
                print(f'error at get_brand_and_name {detail_url}, {type(e).__name__} :: {e}')



    # обычно вся работа с ссылками вынесена в класс URLOperator, но здесь будет рекурсия
    # и желательно сразу отдавать обработанные данные - для получения ссылок требуется обращение к апи
    # пока код будет обращаться к апи, на стороне клиента будет сидеть ждуля в лице водилы

    # в бдшку данные не пишутся, подгружаются в лайв-онли режиме из-за частого обновления и требуемой скорости работы

    # функция возвращает отсортированные, полностью упакованные данные за последние 2 недели
    # в виде списка из словарей
    async def get_supplies(self, token, cabinet, next=0):

        supplies_api_url = f'https://suppliers-api.wildberries.ru/api/v3/supplies?limit=1000&next={next}'
        headers = {'Authorization': f'Bearer {token}'}
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            resp = await client.get(supplies_api_url, timeout=None, headers=HEADERS)
            resp = resp.json()
            supplies_from_resp = resp.get('supplies')
            next = resp.get('next')

            last_supply = supplies_from_resp[-1]
            last_supply_str_date = last_supply.get('createdAt')[:10]
            last_supply_date = datetime.strptime(last_supply_str_date, '%Y-%m-%d').date()

            date_limit = datetime.today().date() - timedelta(weeks=1)

            # если дата последней поставки младше или равна неделе и от апи пришел список менее чем из 1000 поставок
            # то следовательно на сл страницу смысла лезть нет
            # собираем инфу только с этого ответа от апи

            if last_supply_date > date_limit:
                delivery_info = []
                stop_flag = False

                # сортировки нет, переворачиваем список с поставками
                supplies_from_resp.reverse()
                while not stop_flag:

                    # пробегаемся по поставкам, проверям поставки по дате, собираем дату
                    # пополняем delivery_info словариками с датой по поставкам
                    for supply in supplies_from_resp:
                        date_str = supply.get('createdAt')[:10]
                        date = datetime.strptime(date_str, '%Y-%m-%d').date()
                        if date >= date_limit:

                            scan = True if supply.get('scanDt') else False
                            delivery_info.append(
                                {
                                    'qr': supply.get('id'),
                                    'supply_name': supply.get('name'),
                                    'date': date,
                                    'cabinet': cabinet,
                                    'scan': scan
                                }
                            )
                        else:
                            stop_flag = True

                if len(supplies_from_resp) == 1000:
                    return delivery_info + await self.get_supplies(token, cabinet, next)
                return delivery_info

            return await self.get_supplies(token, cabinet, next)

    async def getRequestsData(self):
        proxies = self.proxy_operator.get_random_proxy()

        # получаем ответ от вб с закодированными данными по миллиону топ запросов
        # декодируем текст в человеческий
        async with AsyncClient(proxies=proxies) as client:
            resp = await client.get('https://seller-weekly-report.wildberries.ru/ns/trending-searches/suppliers-portal-analytics/file?period=month', timeout=None, headers=HEADERS)
            print(resp.status_code)
            resp = resp.json()
            enc_data = resp['data']['file']
            data = base64.b64decode(enc_data).decode('utf-8')
            queriesAsStrs = data.split('\n')

            # проходимся по каждой строчке, предварительно сплитили по переносу
            # роспаковываем на запрос и частоту
            # создаем и добавляем словарь с распакованными данными и доп полем normalized_keywords (потребуется для другой ф-и)
            requests_data = []

            for query in queriesAsStrs:
                try:
                    # некоторые запросы содержат неразделительные запятые, например
                    # "постельное 1,5 бязь комплект",40
                    # делаем проверочку, далее создаем, добавляем словарь
                    if query.count(',') > 1:
                        pre_keywords, frequency = query.split('",')
                        keywords = pre_keywords.strip('"')
                    else:
                        keywords, frequency = query.split(',')

                    requests_data.append(
                        {
                            "keywords": keywords.replace('﻿', ''),
                            "frequency": frequency,
                        }
                    )
                except:
                    continue

            return requests_data


    async def get_first_ten_product(self, query_url, phrase):
        nmids = []
        query_url = query_url.replace('replace_me', '1')
        proxies = self.proxy_operator.get_random_proxy()

        async with AsyncClient(proxies=proxies) as client:
            try:
                resp = await client.get(query_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                products = resp.get('data').get('products')
                if not products:
                    return nmids

                counter = 1
                for product in products:
                    if counter <= 10:
                        nmids.append(str(product.get('id', '')))
                        counter += 1
                    else:
                        break

                return nmids
            except Exception as e:
                print(f'{type(e).__name__} :: {e} handled during get_first_ten_products')


    async def get_promotion(self, detail_url):
        try:
            proxies = self.proxy_operator.get_random_proxy()

            async with AsyncClient(proxies=proxies) as client:
                resp = await client.get(detail_url, timeout=None, headers=HEADERS)
                resp = resp.json()
                prom = resp.get('data').get('products')[0].get('promoTextCat')
                return prom
        except Exception as e:
            print(f'{type(e).__name__} :: {e} handled during get_promotion')
            return None


class DataOperator:

    def __init__(self, nmid='', desc=''):
        self.nmid = str(nmid)
        temp_desc = normalize_text(re.sub(r'[^\w\s]', '', desc))
        self.desc = temp_desc.split(' ')


    def check_desc(self, query: dict):

        normalized_keywords = set(query.normalized_keywords.split(' '))
        keywords_lenght = len(query.keywords)


        if normalized_keywords.issubset(set(self.desc)) and keywords_lenght >= 3:
            return True
        return False

    def check_ad(self, ad_ids: list[str]):

        if self.nmid in ad_ids:
            return ad_ids.index(self.nmid) + 1
        return 0


    def check_existence(self, brand_ids):
        if self.nmid in brand_ids:
            return True
        else:
            return False

    def check_first_ten_pages(self, ids):
        if self.nmid in ids:
            return ids.index(self.nmid) + 1

        return 1001


    def check_top_category(self, category_id, subject_base):
        return subject_base.get(category_id)


    def load_and_normalize_request(self, request_data: dict):
        keywords = request_data.get('keywords')
        frequency = request_data.get('frequency')
        normalized_keywords = normalize_text(keywords)
        models.Request.objects.create(
            keywords=keywords, frequency=frequency, normalized_keywords=normalized_keywords
        )

class Indexer:

    url_operator = URLOperator()
    data_collector = DataCollector()
    data_operator = DataOperator()


    def __init__(self, nmid=None):
        self.nmid = nmid

    async def get_standard(self):
        pass
    async def search_common(self, requests_data):

        card_url = self.url_operator.create_card_url(self.nmid)
        detail_card_url = self.url_operator.create_nmid_detail_url(self.nmid)

        card_info = await self.data_collector.get_card_info(card_url)
        detail_info = await self.data_collector.get_brand_and_name(detail_card_url)
        detail_info = ' '.join(detail_info)
        full_info = ' '.join([card_info, detail_info])

        self.data_operator = DataOperator(self.nmid, full_info)
        self.resulted_queries = filter(self.data_operator.check_desc, requests_data)

    async def __get_brand_id(self):
        detail_url = self.url_operator.create_nmid_detail_url(self.nmid)
        brand_id = await self.data_collector.get_brand_id(detail_url)
        return brand_id

    async def get_req_depth(self, query):
        query_req_depth_url = self.url_operator.create_query_req_depth_url(query)
        total = await self.data_collector.get_req_depth(query_req_depth_url)
        return total

    async def __get_existence(self, query):
        brand_id = await self.__get_brand_id()
        query_by_brand_url = self.url_operator.create_filtered_by_brand_url(query=query, brand_id=brand_id)
        brand_products = await self.data_collector.get_query_by_brand(query_by_brand_url)
        self.data_operator.nmid = str(self.nmid)
        existence = self.data_operator.check_existence(brand_products)
        return existence

    async def __get_ad_info(self, query):
        ad_url = self.url_operator.create_ad_url(query)
        ad_products = await self.data_collector.get_ad(ad_url)
        ad_place = self.data_operator.check_ad(ad_products)

        return {
            'ad_spots': len(ad_products),
            'ad_place': ad_place
        }


    async def __get_place(self, query):
        query_url = self.url_operator.create_query_url(query)
        products = await self.data_collector.get_query(query_url)
        place = self.data_operator.check_first_ten_pages(products)
        return place

    async def get_top_category(self, query):
        subject_base_url = self.url_operator.subject_base_url
        ad_info_url = self.url_operator.create_ad_url(query)
        subject_base = await self.data_collector.get_subject_base(subject_base_url)
        top_category_id = await self.data_collector.get_top_category(ad_info_url)
        top_category = self.data_operator.check_top_category(top_category_id, subject_base)

        if top_category:
            return top_category
        else:
            query_category_url = self.url_operator.create_query_categories_url(query)

            most_category = await self.data_collector.get_most_category(query_category_url)
            return most_category

    async def get_prom(self):
        detail_url = self.url_operator.create_nmid_detail_url(self.nmid)
        prom = self.data_collector.get_promotion(detail_url)
        return prom


    def iterate_resulted_queries(self):
        for query in self.resulted_queries:
            keywords = query.keywords
            frequency = query.frequency
            top_category =  async_to_sync(self.get_top_category)(keywords)
            req_depth =  async_to_sync(self.get_req_depth)(keywords)
            existence =  async_to_sync(self.__get_existence)(keywords)

            if existence:
                ad_info = async_to_sync(self.__get_ad_info)(keywords)
                ad_spots = ad_info.get('ad_spots')
                ad_place = ad_info.get('ad_place')
                place =  async_to_sync(self.__get_place)(keywords)
                # prom = async_to_sync(self.get_prom)()

                if req_depth != 0:
                    if not place:
                        place = req_depth + 1
                    percent = (place / req_depth) * 100
                    spot_req_depth = str(round(percent, 2)).replace('.', ';')
            else:
                ad_spots, ad_place, place, spot_req_depth = [None] * 4

            yield {
                'nmid': self.nmid, 'top_category': top_category,
                'keywords': keywords, 'frequency': frequency,
                'req_depth': req_depth, 'existence': existence,
                'place': place, 'spot_req_depth': spot_req_depth,
                'ad_spots': ad_spots, 'ad_place': ad_place
            }

    def iterate_standard_queries(self, standard_queries):
        for query in standard_queries:
            keywords = query.query
            frequency = query.frequency
            top_category =  async_to_sync(self.get_top_category)(keywords)
            req_depth =  async_to_sync(self.get_req_depth)(keywords)
            existence =  async_to_sync(self.__get_existence)(keywords)

            if existence:
                ad_info = async_to_sync(self.__get_ad_info)(keywords)
                ad_spots = ad_info.get('ad_spots')
                ad_place = ad_info.get('ad_place')
                place =  async_to_sync(self.__get_place)(keywords)

                if req_depth != 0:
                    if not place:
                        place = req_depth + 1
                    percent = (place / req_depth) * 100
                    spot_req_depth = str(round(percent, 2)).replace('.', ';')
            else:
                ad_spots, ad_place, place, spot_req_depth = [None] * 4

            yield {
                'nmid': self.nmid, 'top_category': top_category,
                'keywords': keywords, 'frequency': frequency,
                'req_depth': req_depth, 'existence': existence,
                'place': place, 'spot_req_depth': spot_req_depth,
                'ad_spots': ad_spots, 'ad_place': ad_place
            }



    def iterate_resulted_queries_seo(self):
        for query in self.resulted_queries:
            keywords = query.keywords
            depth = async_to_sync(self.get_req_depth)(keywords)
            top_category = async_to_sync(self.get_top_category)(keywords)
            yield {
                'keywords': keywords, 'top_category': top_category,
                'frequency': query.frequency, 'depth': depth

            }

class SeoCollector:

    url_operator = URLOperator()
    data_collector = DataCollector()
    indexer = Indexer()
    depth = 10

    def __init__(self, phrase):
        self.requests_data = list(models.Request.objects.all())
        self.phrase = phrase


    def run(self):
        if models.SeoCollectorPhrase.objects.filter(phrase=self.phrase).first() is None:
            req_depth = async_to_sync(self.indexer.get_req_depth)(self.phrase)
            priority_cat = async_to_sync(self.indexer.get_top_category)(self.phrase)
            phrase_obj = models.SeoCollectorPhrase.objects.create(phrase=self.phrase, priority_cat=priority_cat, req_depth=req_depth)

            query_url = self.url_operator.create_query_url(self.phrase)
            products = async_to_sync(self.data_collector.get_first_ten_product)(query_url, self.phrase)
            seen = set()

            for product in products:
                indexer = Indexer(product)
                async_to_sync(indexer.search_common)(self.requests_data)

                for query in indexer.iterate_resulted_queries_seo():
                    keywords = query.get('keywords')
                    if not keywords in seen:
                        if query.get('top_category') == phrase_obj.priority_cat:
                            standard = True
                        else:
                            standard = False
                        models.SeoCollectorPhraseData.objects.create(
                            query=keywords, priority_cat=query.get('top_category'),
                            phrase=phrase_obj, standard=standard,
                            depth=query.get('depth'), frequency=query.get('frequency')
                        )
                    seen.add(keywords)

            phrase_obj.ready = True
            phrase_obj.save()


#################### SCRAPER PHRASES #####################
async def get_most_categories(phrase):
    category = ''
    count = 0
    categories_url = 'https://search.wb.ru/exactmatch/ru/common/v4/search?TestGroup=test_2&TestID=131&' \
                    'appType=1&curr=rub&dest=123586150&filters=xsubject&query={}&regions=80,38,4,64,83' \
                    ',33,68,70,69,30,86,40,1,66,110,22,31,48,71,114&resultset=filters&spp=0'.format(phrase)
    async with AsyncClient(proxies=proxies) as client:
        try:
            resp = await client.get(categories_url, timeout=None, headers=HEADERS)
            resp = resp.json()
            categories = resp.get('data').get('filters')[0].get('items')
            categories.sort(key=lambda category: int(category.get('count')) or 0)

            while len(categories) < 3:
                categories.append({})

            return list(map(lambda x: x.get('name') or '', categories))[:3]

        except Exception as e:
            print(f'error at  get_query_most_categories {phrase} {e}')
            return ['', '', '']


class ScraperPhrases:
    scraper = DataCollector()
    url_operator = URLOperator()
    data_operator = DataOperator()

    @classmethod
    def initializate_subject_base(cls):
        subject_base_url = cls.url_operator.subject_base_url
        cls.subject_base = async_to_sync(cls.scraper.get_subject_base)(subject_base_url)

    @classmethod
    def scraping_phrase(cls, phrase):
        ad_info_url = cls.url_operator.create_ad_url(phrase)
        prior_category_id = async_to_sync(cls.scraper.get_top_category)(ad_info_url)
        prior_catogory = cls.data_operator.check_top_category(prior_category_id, cls.subject_base)
        req_depth = async_to_sync(cls.scraper.get_req_depth)(cls.url_operator.create_query_req_depth_url(phrase))

        top_category, second_top_category, third_top_category = async_to_sync(get_most_categories)(phrase)
        return prior_catogory, req_depth, top_category, second_top_category, third_top_category


def get_filter_and_sorted_context(request):
    sorted_context = []
    filter_context = {}
    sorted_by = request.GET.get("sorted_by")
    category = request.GET.get("category")
    top_category = request.GET.get("top_category")
    search_phrase = request.GET.get("search_phrase")
    search_category = request.GET.get("search_category")
    search_top_category = request.GET.get("search_top_category")

    if sorted_by is not None:
        sorted_context.extend(sorted_by.replace(" ", "").split(','))

    if category is not None:
        filter_context['priority_cat'] = category

    if search_phrase is not None:
        filter_context['phrase__icontains'] = search_phrase

    if search_category is not None:
        filter_context['priority_cat__icontains'] = search_category

    if search_top_category is not None:
        filter_context['top_category__icontains'] = search_top_category

    if top_category is not None:
        filter_context['top_category'] = top_category

    return sorted_context, filter_context


def create_xlsx_table(queryset):
    buffer = io.BytesIO()
    book = xlsxwriter.Workbook(buffer)
    sheet = book.add_worksheet()

    columns = [
        'Ключевые слова', "Частотность", "Глубина", "Рекламная категория",
        "Топовая категория", "Вторая категория", "Третья категория",
    ]

    for col_counter, column in enumerate(columns):
        sheet.write(0, col_counter, column)

    sheet.set_column(1, 3, 25)
    sheet.set_column(0, 0, 100)

    for row_counter, phrase in enumerate(queryset, start=1):
        print(row_counter)
        sheet.write(row_counter, 0, phrase.phrase)
        sheet.write(row_counter, 1, phrase.frequency)
        sheet.write(row_counter, 2, phrase.req_depth)
        sheet.write(row_counter, 3, phrase.priority_cat)
        sheet.write(row_counter, 4, phrase.top_category)
        sheet.write(row_counter, 5, phrase.second_top_category)
        sheet.write(row_counter, 6, phrase.third_top_category)

    book.close()
    buffer.seek(0)
    return buffer
